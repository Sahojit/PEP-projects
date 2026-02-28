import os
from datetime import datetime
from fastapi import FastAPI
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "data.xlsx")

print("=== FASTAPI STARTUP ===")
print("BASE_DIR:", BASE_DIR)
print("EXCEL_PATH:", EXCEL_PATH)

# ==============================
# App Setup
# ==============================
app = FastAPI()

# ==============================
# Create Excel if not exists
# ==============================
if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    sheet = wb.active
    sheet.append(["ID", "Name", "Age", "City", "Timestamp"])
    wb.save(EXCEL_PATH)
    wb.close()
    print("Created new Excel file at:", EXCEL_PATH)
else:
    print("Using existing Excel file at:", EXCEL_PATH)

# ==============================
# Request Model
# ==============================
class Person(BaseModel):
    name: str
    age: int
    city: str

# ==============================
# Routes
# ==============================

@app.get("/")
def health():
    return {
        "status": "FastAPI Excel Auto-Update Server Running",
        "excel_file": EXCEL_PATH
    }

@app.post("/add")
def add_person(person: Person):
    print("\n=== ADD PERSON CALLED ===")
    print("Writing to Excel file at:", EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active

    before_rows = sheet.max_row
    print("Rows BEFORE append:", before_rows)

    next_id = before_rows
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    sheet.append([
        next_id,
        person.name,
        person.age,
        person.city,
        timestamp
    ])

    after_rows = sheet.max_row
    print("Rows AFTER append:", after_rows)

    wb.save(EXCEL_PATH)
    wb.close()

    # Double-check by reloading
    wb2 = load_workbook(EXCEL_PATH)
    sheet2 = wb2.active
    verify_rows = sheet2.max_row
    wb2.close()

    print("Rows AFTER RELOAD:", verify_rows)

    return {
        "status": "success",
        "excel_file": EXCEL_PATH,
        "rows_before": before_rows,
        "rows_after": after_rows,
        "rows_after_reload": verify_rows,
        "added_row": {
            "id": next_id,
            "name": person.name,
            "age": person.age,
            "city": person.city,
            "timestamp": timestamp
        }
    }

@app.get("/all")
def get_all_data():
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active

    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        records.append({
            "id": row[0],
            "name": row[1],
            "age": row[2],
            "city": row[3],
            "timestamp": row[4]
        })

    wb.close()

    return {
        "excel_file": EXCEL_PATH,
        "total_rows": len(records),
        "data": records
    }
